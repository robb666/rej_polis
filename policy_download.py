from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.options import Options
import time
from openpyxl import load_workbook
import os
import io
import re
import win32com.client
from win32com.client import Dispatch
import datetime


wb = load_workbook(filename="M:\\Agent baza\Login_Hasło.xlsx", read_only=True)
ws = wb['Arkusz1']
tuz_l = ws['F57'].value
tuz_h = ws['G57'].value


"""CHROME"""
n = int(input('Wpisz ilość polis do zarejestrowania: '))


def chrome_ustawienia():
    """Chrome ustawienia"""
    options = webdriver.ChromeOptions()
    preferences = {'download.default_directory': os.getcwd() + "tuz_polisy"}
    options.add_experimental_option("prefs", preferences)
    driver = webdriver.Chrome(executable_path=r'M:\zzzProjekty\drivery przegądarek\chromedriver.exe', options=options)

    return driver


def tuz_logowanie(driver):
    """Logowanie"""
    url_tuz = 'https://sobol-agencyjny.tuz.pl/'
    driver.get(url_tuz)
    login = driver.find_element_by_name('user_login')
    login.send_keys(tuz_l)
    haslo = driver.find_element_by_name('user_password')
    haslo.send_keys(tuz_h)
    driver.find_element_by_css_selector('.form-submit').click()


def lista_polis():
    """Przejście do listy polis Marka"""
    driver.get('https://sobol-agencyjny.tuz.pl/insured/contracts')
    driver.find_element_by_id('contract_type').click()
    driver.find_element_by_css_selector('#contract_type > option:nth-child(4)').click()
    driver.find_element_by_id('register_name').send_keys('wołowski')
    driver.find_element_by_id('search_handler').click()


def kolejna_polisa():
    """Klika w dane każdej polisy"""
    try:
        for i in range(n, 0, -1):
            if 50 < i <= 75:
                WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, 'contracts_next'))).click()
                WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.ID, 'contracts_next'))).click()
                ccc = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#contracts > tbody > tr:nth-child(' + str(i - 50) + ') > td:nth-child(8) > a:nth-child(1) > input')))
                ccc.click()
                yield driver

            elif 25 < i <= 50:
                try:
                    WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, 'contracts_next'))).click()
                    cc = WebDriverWait(driver, 9).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#contracts > tbody > tr:nth-child(' + str(i - 25) + ') > td:nth-child(8) > a:nth-child(1) > input')))
                    cc.click()
                except:
                    imie_nazw = driver.find_element_by_css_selector('#contracts > tbody > tr:nth-child(' + str(i - 25) + ') > td:nth-child(4)').text
                    ser = driver.find_element_by_css_selector('#contracts > tbody > tr:nth-child(' + str(i - 25) + ') > td:nth-child(1)').text
                    nr = driver.find_element_by_css_selector('#contracts > tbody > tr:nth-child(' + str(i - 25) + ') > td:nth-child(2)').text
                    WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, 'contracts_previous'))).click()
                    print(f'NIE Zapisał {imie_nazw} {ser}{nr}')
                    # pass
                yield driver

            elif 0 < i:
                c = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#contracts > tbody > tr:nth-child(' + str(i) + ') > td:nth-child(8) > a:nth-child(1) > input')))
                c.click()
                yield driver


    except Exception as err:
        print(err)


def szukanie_danych():
    """"""
    for _ in kolejna_polisa():
        nr_polisy = ''
        try:
            seria_polisy = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#main > div > div.mybox > h1 > table > tbody > tr > td:nth-child(1) > nobr'))).text
            polisa_nr = driver.find_element_by_css_selector('#main > div > div.mybox > h1 > table > tbody > tr > td:nth-child(1) > nobr').text
            nr_polisy = seria_polisy[-11:-8] + polisa_nr[-7:]

            nr_polisy_wzn = ''
            try:
                seria_polisy_wzn = driver.find_element_by_css_selector('#tabs-packages > fieldset > fieldset.group_qual.group_qual_legend > table > tbody > tr > td:nth-child(2)').text
                polisa_nr_wzn = driver.find_element_by_css_selector('#tabs-packages > fieldset > fieldset.group_qual.group_qual_legend > table > tbody > tr > td:nth-child(4)').text
                nr_polisy_wzn = seria_polisy_wzn + polisa_nr_wzn
            except:
                pass

            data_zawarcia = ''
            if 'KOS' in seria_polisy:
                try:
                    data_zawarcia = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#tabs-packages > fieldset > fieldset.group_qual.fieldset_noborder > table > tbody > tr:nth-child(1) > td:nth-child(2)'))).text
                    # data_zawarcia = driver.find_element_by_css_selector('#tabs-packages > fieldset > fieldset.group_qual.fieldset_noborder > table > tbody > tr:nth-child(1) > td:nth-child(2)').text
                    data_zawarcia = datetime.datetime.strptime(data_zawarcia[2:], '%y-%m-%d')
                except:
                    pass

            else:
                try:
                    data_zawarcia = driver.find_element_by_css_selector('#tabs-packages > fieldset > fieldset.group_qual.fieldset_noborder > table > tbody > tr:nth-child(1) > td:nth-child(4)').text
                    data_zawarcia = datetime.datetime.strptime(data_zawarcia[2:], '%y-%m-%d')
                except:
                    pass

            nowa_wzn_kos = driver.find_element_by_css_selector('#tabs-packages > fieldset > fieldset.group_qual.fieldset_noborder > table > tbody > tr:nth-child(1) > td:nth-child(4)').text
            nowa_wzn_brs = driver.find_element_by_css_selector('#tabs-packages > fieldset > fieldset.group_qual.fieldset_noborder > table > tbody > tr:nth-child(1) > td:nth-child(2)').text
            nowa_wzn = 'W' if 'wznowienie' in nowa_wzn_kos.lower() or 'wznowienie' in nowa_wzn_brs.lower() else 'N'

            driver.find_element_by_id('ui-id-2').click()

            nazwisko = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.group_qual.group_qual_legend > table > tbody > tr:nth-child(2) > td:nth-child(4)').text
            if 'KOS' in seria_polisy:
                imie = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.group_qual.group_qual_legend > table > tbody > tr:nth-child(1) > td:nth-child(4)').text
            else:
                imie = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.group_qual.group_qual_legend > table > tbody > tr:nth-child(2) > td:nth-child(2)').text

            pesel = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.group_qual.group_qual_legend > table > tbody > tr:nth-child(1) > td:nth-child(2)').text
            data_prawka = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.group_qual.group_qual_legend > table > tbody > tr:nth-child(2) > td:nth-child(2)').text
            data_pr_j = data_prawka if 'KOS' in seria_polisy else ''

            ulica = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.customer_address_type_live.group_qual.group_is_show_1 > table > tbody > tr:nth-child(2) > td:nth-child(2)').text
            ulica1 = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.customer_address_type_live.group_qual.group_is_show_1 > table > tbody > tr:nth-child(2) > td:nth-child(4)').text
            ulica = ulica if ulica != 'ŁÓDZKIE' else ulica1
            nr_ul = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.customer_address_type_live.group_qual.group_is_show_1 > table > tbody > tr:nth-child(3) > td:nth-child(2)').text
            nr_ul1 = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.customer_address_type_live.group_qual.group_is_show_1 > table > tbody > tr:nth-child(3) > td:nth-child(4)').text
            nr_ul = nr_ul if nr_ul != 'ZGIERSKI' else nr_ul1

            kod_poczt = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.customer_address_type_live.group_qual.group_is_show_1 > table > tbody > tr:nth-child(1) > td:nth-child(2)').text
            miasto = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.customer_address_type_live.group_qual.group_is_show_1 > table > tbody > tr:nth-child(4) > td:nth-child(2)').text
            miasto1 = driver.find_element_by_css_selector('#clone_customer_0_ > fieldset.customer_address_type_live.group_qual.group_is_show_1 > table > tbody > tr:nth-child(4) > td:nth-child(4)').text
            nr_m = ''
            if re.search('[0-9]', miasto):
                nr_m = 'm ' + miasto
                miasto = miasto1

            adres = f'{ulica} {nr_ul} {nr_m}'

            driver.find_element_by_id('ui-id-3').click()

            marka = driver.find_element_by_css_selector('#tabs-objects > div > fieldset > fieldset > table > tbody > tr:nth-child(4) > td:nth-child(2)').text
            model = driver.find_element_by_css_selector('#tabs-objects > div > fieldset > fieldset > table > tbody > tr:nth-child(5) > td:nth-child(2)').text
            nr_rej = driver.find_element_by_css_selector('#tabs-objects > div > fieldset > fieldset > table > tbody > tr:nth-child(3) > td:nth-child(4)').text
            rok_prod = driver.find_element_by_css_selector('#tabs-objects > div > fieldset > fieldset > table > tbody > tr:nth-child(3) > td:nth-child(2)').text

            driver.find_element_by_id('ui-id-4').click()
            data_pocz = ''
            data_konca = ''
            try:
                data_pocz = driver.find_element_by_css_selector('#clone_productobject_55381_179227_ > div > fieldset > fieldset > fieldset > fieldset > table > tbody > tr:nth-child(2) > td:nth-child(2)').text
                data_pocz = datetime.datetime.strptime(data_pocz[2:], '%y-%m-%d')
                data_konca = driver.find_element_by_css_selector('#clone_productobject_55381_179227_ > div > fieldset > fieldset > fieldset > fieldset > table > tbody > tr:nth-child(2) > td:nth-child(4)').text
                data_konca = datetime.datetime.strptime(data_konca[2:], '%y-%m-%d')
            except:
                driver.find_element_by_id('ui-id-1').click()
                try:
                    data_pocz = driver.find_element_by_css_selector('#tabs-packages > fieldset > fieldset.group_qual.fieldset_noborder > table > tbody > tr:nth-child(2) > td:nth-child(2)').text
                    data_pocz = datetime.datetime.strptime(data_pocz[2:], '%y-%m-%d')
                except:
                    pass
                try:
                    data_konca = driver.find_element_by_css_selector('#tabs-packages > fieldset > fieldset.group_qual.fieldset_noborder > table > tbody > tr:nth-child(2) > td:nth-child(4)').text
                    data_konca = datetime.datetime.strptime(data_konca[2:], '%y-%m-%d')
                except:
                    pass

            driver.find_element_by_id('ui-id-4').click()

            rodzaj = 'kom' if 'KOS' in seria_polisy else 'rol'

            driver.find_element_by_id('ui-id-5').click()

            przypis = driver.find_element_by_css_selector('#tabs-tariff > fieldset > fieldset:nth-child(2) > table > tbody > tr > td:nth-child(2)').text
            tel = ''
            tel_szukaj = driver.find_elements_by_css_selector('#tabs-tariff > fieldset > fieldset.group_qual.fieldset_noborder > table > tbody ')
            for i in tel_szukaj:
                tel = i.text.split('\n')[-1]

            ter_platnosci = ''
            try:
                ter_platnosci = driver.find_element_by_css_selector('#tabs-tariff > fieldset > fieldset:nth-child(2) > table > tbody > tr > td:nth-child(4)').text
                ter_platnosci = datetime.datetime.strptime(ter_platnosci[2:], '%y-%m-%d')
            except:
                pass

            p_czy_g = 'P' if 'Przelew' in driver.page_source else 'G'
            ilosc_rat = '1' if 'JEDNORAZOWA' in driver.page_source or 'jednorazowej' in driver.page_source else ''
            nr_raty = '1' if ilosc_rat else ''

            driver.execute_script("window.history.go(-1)")
            driver.find_element_by_id('search_handler').click()

        except:
            # print(f'NIE Zapisał {nr_polisy}')
            nazwisko, imie, pesel, data_pr_j, adres, kod_poczt, miasto, tel, marka, seria_polisy, \
            model, nr_rej, rok_prod, data_zawarcia, data_pocz, data_konca, rodzaj, nr_polisy, nowa_wzn, \
            nr_polisy_wzn, przypis, ter_platnosci, p_czy_g, nr_raty, ilosc_rat = \
                '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',



        """Zapisanie w Bazie"""
        path = os.getcwd()

        # Sprawdza czy arkusz jest otwarty
        try:
            ExcelApp = win32com.client.GetActiveObject('Excel.Application')
            wb = ExcelApp.Workbooks(path + "\\marek_tuz.xlsx")
            # ws = wb.Worksheets("Arkusz1")
            # workbook = ExcelApp.Workbooks("Baza.xlsx")

        # Jeżeli arkusz jest zamknięty, otwiera go
        except:
            ExcelApp = Dispatch("Excel.Application")
            wb = ExcelApp.Workbooks.Open(path + "\\marek_tuz.xlsx")
            # ws = wb.Worksheets("Arkusz1")

        row_to_write = wb.Worksheets(1).Cells(wb.Worksheets(1).Rows.Count, 12).End(-4162).Row + 1

        ExcelApp.Cells(row_to_write, 7).Value = 'Marek'
        ExcelApp.Cells(row_to_write, 10).Value = 'Wołowski'
        # ExcelApp.Cells(row_to_write, 11).Value = firma
        ExcelApp.Cells(row_to_write, 12).Value = nazwisko #if nazwisko else print(f'NIE Zapisał {ser}{nr}')
        ExcelApp.Cells(row_to_write, 13).Value = imie
        ExcelApp.Cells(row_to_write, 14).Value = 'p' + pesel
        ExcelApp.Cells(row_to_write, 15).Value = data_pr_j
        ExcelApp.Cells(row_to_write, 16).Value = adres
        ExcelApp.Cells(row_to_write, 17).Value = kod_poczt
        ExcelApp.Cells(row_to_write, 18).Value = miasto
        ExcelApp.Cells(row_to_write, 19).Value = tel if not re.search('[A-z]', tel) else ''
        # ExcelApp.Cells(row_to_write, 20).Value = email
        ExcelApp.Cells(row_to_write, 23).Value = marka if 'KOS' in seria_polisy else ''
        ExcelApp.Cells(row_to_write, 24).Value = model if 'KOS' in seria_polisy else ''
        ExcelApp.Cells(row_to_write, 25).Value = nr_rej if 'KOS' in seria_polisy else ''
        ExcelApp.Cells(row_to_write, 26).Value = rok_prod if 'KOS' in seria_polisy else ''
        # ExcelApp.Cells(row_to_write, 29).Value = int(ile_dni) + 1
        ExcelApp.Cells(row_to_write, 30).Value = data_zawarcia
        ExcelApp.Cells(row_to_write, 31).Value = data_pocz
        ExcelApp.Cells(row_to_write, 32).Value = data_konca
        ExcelApp.Cells(row_to_write, 36).Value = 'SPÓŁKA'
        ExcelApp.Cells(row_to_write, 37).Value = 'TUZ'
        ExcelApp.Cells(row_to_write, 38).Value = 'TUZ'
        ExcelApp.Cells(row_to_write, 39).Value = rodzaj
        ExcelApp.Cells(row_to_write, 40).Value = nr_polisy
        ExcelApp.Cells(row_to_write, 41).Value = nowa_wzn
        ExcelApp.Cells(row_to_write, 42).Value = nr_polisy_wzn
        # ryzyko = ExcelApp.Cells(row_to_write, 46).Value = 'b/d'
        ExcelApp.Cells(row_to_write, 48).Value = przypis.strip(' PLN')
        ExcelApp.Cells(row_to_write, 49).Value = ter_platnosci
        ExcelApp.Cells(row_to_write, 50).Value = przypis.strip(' PLN')
        ExcelApp.Cells(row_to_write, 51).Value = p_czy_g
        ExcelApp.Cells(row_to_write, 52).Value = nr_raty
        ExcelApp.Cells(row_to_write, 53).Value = ilosc_rat
        ExcelApp.Cells(row_to_write, 54).Value = ter_platnosci
        ExcelApp.Cells(row_to_write, 55).Value = przypis.strip(' PLN')
        ExcelApp.Cells(row_to_write, 59).Value = 'TUZ'

        wb.Save()
        wb.Close()
        if nazwisko:
            print(f'Zapisał {nazwisko} {nr_polisy}')



try:
    driver = chrome_ustawienia()
    tuz_logowanie(driver)
    lista_polis()
    kolejna_polisa()
    szukanie_danych()

except:
    print('Błąd')
    pass
time.sleep(9000)
